import openpyxl as xl
from datetime import date, time, datetime, timedelta
import math

my_castle_x = 647
my_castle_y = 645

def createSpreadsheet():
    #wb = xl.load_workbook('desert_fortresses.xlsx')
    wb = xl.Workbook()
    ws = wb.active
    ws.append(["coord X", "coord Y", "hours", "minutes", "seconds", "free in"])

def updateSpreadsheet(coord_x, coord_y, hours = 0, minutes = 0, seconds = 0, reign = 'desert', lvl = 0):
    if(reign == 'desert'):
        wb = xl.load_workbook('desert_fortresses.xlsx')
    if(reign == 'peak'):
        wb = xl.load_workbook('peak_fortresses.xlsx')
    if(reign == 'islands'):
        wb = xl.load_workbook('storm_islands.xlsx')

    ws = wb.active
    if(reign == 'peak' or reign == 'desert'):
        free_in = datetime.today() + timedelta(hours = hours, minutes = minutes, seconds = seconds)
        ws.append([coord_x, coord_y, free_in.strftime('%H:%M:%S')])
    else:
        ws.append([coord_x, coord_y, lvl, math.dist([coord_x, coord_y], [my_castle_x, my_castle_y])])

    if(reign == 'desert'):
        wb.save('desert_fortresses.xlsx')
    if(reign == 'peak'):
        wb.save('peak_fortresses.xlsx')
    if(reign == 'islands'):
        wb.save('storm_islands.xlsx')